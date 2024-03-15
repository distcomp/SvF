# -*- coding: cp1251 -*-

#from   numpy import *
import   numpy as np
from   os.path  import *
import openpyxl
import sys
#import COMMON as co
from   Pars   import *
##from   InData import *
from Parser    import *
from Object    import *

#def  TblFldV ( TblFld ) : return getTbl_tbl ( TblFld )
#def  Tbl     ( nTbl ) :   return getTbl     ( nTbl   )

def ParseSelect30(buf):  ##  разбор  Select
    buf = buf.strip().replace(', ',",")
#    print (buf)
    part = SplitIgnor(buf, 'Select ')
    #      print (part)
    leftName = part[0][:-1]
    part = SplitIgnor(part[1], ' from ')
    #       print ('part_from',part)
    Fields = part[0]
    part = SplitIgnor(part[1], ' where ')
    if len(part) == 1:
        where = ''
    else:
        where = part[1]
    part = SplitIgnor(part[0], ' As ')
    FileName = part[0]
    if len(part) == 2:
        AsName = part[1]
    else:
        AsName = ''
    if AsName == '':  AsName = leftName
    #       print (leftName,'!', Fields, '!',FileName,'!'+AsName+'!'+where+'!' )
    return leftName, Fields, FileName, AsName, where


def  Treat_FieldNames (raw_line) :
        for itb, tb in enumerate ( SvF.Task.Objects ) :  #  from Tbls   Dat.X -> Dat.dat('X')
            if tb.Otype == 'Table' :
 #               tb.Oprint()
 #               print ('tb.name',tb.name)
                if findNamePos(raw_line, tb.name ) < 0 : continue
                print('\nTABB ' + raw_line, tb.name)
                pars = parser (raw_line)
                itn = 0
                while True :
                    itn = pars.find_part(tb.name, itn)
                    if itn < 0 : break;
                    if itn+2 >= len (pars.items) : break
                    if pars.items[itn+1].part != '.' : break
                    fld = pars.items[itn+2].part
 #                   print ('fld', fld)
                    if ( fld !='dat' and fld !='AddField' and fld !='sR' and fld !='NoR'  and
                         fld !='AppendRec' and fld !='AppendRec1' and
                         fld != 'WriteSvFtbl' and fld != 'Flds' and fld != 'KillRowsByMask' and fld != 'KillField'
                       ):
                        pars.items[itn+2].part = 'dat(\''+fld+'\')'
 #                       raw_line = SubstitudeName(raw_line, fld, 'dat(\''+fld+'\')')
  #                      print ('\nTABA '+pars.join(), pars.items[itn].part, pars.items[itn+1].part, pars.items[itn+2].part)
  #                      print('\nTABE ' + raw_line)
                    itn += 3
                raw_line = pars.join()
#                print('\nTABA ' + pars.join() )
        return raw_line


def joinTab(*Tabs):
        for itbl, tbl in enumerate(Tabs):
            if itbl == 0:
                ret = deepcopy(tbl)
                ret.name = 'joinTab_Result'
            else:
                for fld in tbl.Flds:
                    if ret.getFieldNum(fld.name) == -1:
                        ret.Flds.append(deepcopy(fld))
            print(tbl.NoR)

## 30       SvF.Task.KillTbl(ret.name)  # kill the same name
##        SvF.Task.AddTbl(ret)
        ret.Add()
        SvF.curentTabl = ret
        return ret


def joinTabBy(By, *Tabs):                #  TMo = joinTabBy ( 't', TMos, Spline )
    for j, tbl in enumerate(Tabs):
        if j == 0:
 #           print (tbl.NoR)
            ret = deepcopy(tbl)
            ret.name = 'joinTabBy_Result'
            retByNum = ret.getFieldNum(By)
            if retByNum == -1:
                print ( 'Err find Field by', By )
                exit (-1)
        else:
            tbj = deepcopy(tbl)
            tbjByNum = tbj.getFieldNum(By)
            if tbjByNum == -1:
                print ( 'Err find Field by', By )
                exit (-1)

            r = -1
  #          print ('shape', ret.Flds[retByNum].tb.shape[0])
            while ( r < ret.NoR-1 and r < tbj.NoR-1 ):
                r +=1
   #             print (r, ret.Flds[retByNum].tb[r] , tbj.Flds[tbjByNum].tb[r])
                if ret.Flds[retByNum].tb[r] == tbj.Flds[tbjByNum].tb[r] : continue
                ind = where (tbj.Flds[tbjByNum].tb[r:]==ret.Flds[retByNum].tb[r])
#                print (ind[0])
 #               print (len(ind[0]))
  #              print (ind[0].shape)
                if ind[0].size == 0 :
                    ret.KillRow (r)
                    r -= 1
                    continue
                else :
                    tbj.KillRow (r)
                    r -= 1
                    continue

            while ret.NoR != tbj.NoR :
                if ret.NoR > tbj.NoR : ret.KillRow (ret.NoR-1)
                else                 : tbj.KillRow (tbj.NoR-1)
    #        print('NoR RET TBJ', ret.NoR, tbj.NoR)
            ret = joinTab (ret, tbj)
## 30    SvF.Task.KillTbl(ret.name)  # kill the same name
##    SvF.Task.AddTbl(ret)
    ret.Add()
    SvF.curentTabl = ret
    return ret


def appendTab(*Tabs):
        for itbl, tbl in enumerate(Tabs):
            if itbl == 0:
                ret = deepcopy(tbl)
                ret.name = 'appendTab_Result'
            else:
                for ifld, rfld in enumerate(ret.Flds):
                    rfld.tb = append (rfld.tb, tbl.Flds[ifld].tb)
        ret.NoR = ret.Flds[0].tb.size
        self.sR = range (self.NoR)
## 30        SvF.Task.KillTbl(ret.name)  # kill the same name
##        SvF.Task.AddTbl(ret)
        ret.Add()
        SvF.curentTabl = ret
        return ret


def  TblOperation (buf):
        if buf == '' : return
##30        buf = SvF.Task.substitudeDef(buf)
        buf = buf.replace(' ', '') 
 #       print  ('buf=', buf)
        parts = buf.split ( '=TblLatLonToAzimut(' )             #  LatLonToAzimut
        if len(parts)==1 : parts = buf.split ( '=TblLatLonToGaussKruger(' )    # LanLonToGaussKruger
        if len(parts)==2 :
            parts[1] = parts[1].split(')')[0]
            if SvF.printL: print (parts)
            lefts = parts[0].split(',')
            x_tbl = getTbl_tbl (lefts[0])
            y_tbl = getTbl_tbl (lefts[1])
            rights = parts[1].split(',')
            lat_tbl = getTbl_tbl (rights[0])
            lon_tbl = getTbl_tbl (rights[1])
     #       print (lefts, rights)
            if buf.find ('TblLatLonToAzimut') > 0 :
                for i in range(x_tbl.shape[0]):  x_tbl[i],y_tbl[i] = LatLonToAzimut( lat_tbl[i],lon_tbl[i] )
            else :
                for i in range(x_tbl.shape[0]):  x_tbl[i],y_tbl[i] = WGS84toGausKru( lat_tbl[i],lon_tbl[i],0 )
            return

        parts = buf.split ( '.AddField(' )             #  AddField
        if len(parts)==2 :
            parts[1] = parts[1].split(')')[0]
            getTbl(parts[0]).AddField( parts[1] )
      #      print ('AddField :', parts[1], 'to', parts[0])
            return
        
        parts = buf.replace ( ' ', '').split ( '=' )
        left  = parts[0]
        right = parts[1]
        for itb, tb in enumerate ( SvF.Task.Tbls ) :
#            print 'tb.name', tb.name
            for ifld, fld in enumerate (tb.Flds) :
                right = right.replace ( tb.name +'.'+ fld.name,
                                        'SvF.Task.Tbls['+str(itb)+'].Flds['+str(ifld)+'].tb')
  #      print (right)

        tb_num, fld_num = getTblFieldNums (left)
        SvF.Task.Tbls[tb_num].Flds[fld_num].tb = eval (right)
   #     print ('Operation', parts)
        return
       

class Field :
    def __init__ ( self, name, src_name ): 
        self.name     = name
        self.tb       = None
        self.src_name = src_name
        self.src_num  = NaN

    def Mprint ( self ) :  print ('Field :', self.name, self.src_name, self.src_num)


class Table (Object):
    def __init__ ( self, fromFile, AsName='', fields='*', where_condition=None ): #
        Object.__init__( self, AsName, 'Table')
        if SvF.Compile :  return
        self.fromFile = fromFile
        self.FileType  = 'tbl'
        self.FileVer   = 0
        self.Flds      = []
        self.NoC     = 0
        self.NoR     = 0
        self.sR      = []
        self.useNaN  = SvF.useNaN
        self.fields_str  = fields #.strip()
        self.con_list = []                            # list of arguments for
        self.where_condition = where_condition        #   where_condition

        if self.name == '' : self.name = 'curentTabl'
        if SvF.printL : print ('\nSelect', fields, '\n  from', fromFile, 'name:'+self.name+'|')
#        print('*Table*', fields, ' from', fromFile, 'name:'+self.name+'|')
        ff_nn = SplitIgnor ( fromFile, ' AS ' )     # Имя файла и таблицы

        _fields = self.fields_str.split(',')                   # Fields
        if SvF.printL : print (_fields)
        for fld in _fields :
            part = SplitIgnor ( fld.strip(), ' AS ' )
            src_name = part[0]
            if len(part) == 2 : name = part[1]              # As
            else              : name = part[0]              # the same name
            self.Flds.append ( Field ( name, src_name ) )
            if src_name.upper() == 'ROWNUM' :  self.Flds[-1].src_num = -1
            if SvF.printL : self.Flds[-1].Mprint()

        root, ext = splitext(self.fromFile.upper())
#        if   getTblNum(self.fromFile) !=-1 :  self.Read30_TBL ( )
        if   not getTbl(self.fromFile) is None :  self.Read30_TBL ( )
#        elif not getFun(self.fromFile) is None :  self.Read28_FUN ( where_condition )  # надо переписать
        elif not getFun(self.fromFile) is None :  self.Read28_FUN ( )  # надо переписать
        else :
            self.fromFile = com.DataPath + self.fromFile
            if     '.XLSX' == ext :                   self.Read30_XLSX( )
            elif   '.KML'  == ext :                   self.Read21_KML ( where_condition )  # надо переписать
            elif   '.ASC'  == ext :                   self.Read27_ASC ( where_condition )  # надо переписать
            else :                                    self.Read30_TXT ( )

        print (self.name, ' NoR =', self.NoR, '\n')
        if SvF.printL :
            for ifld, fld in enumerate (self.Flds) :
                print ('Field :', fld.name, fld.src_name, fld.src_num, fld.tb.min(0), fld.tb.max(0))

        self.sR = range (self.NoR)


## 30        SvF.Task.KillTbl ( self.name )   # kill the same name
## 30        SvF.Task.AddTbl ( self )
        SvF.curentTabl = self

    def Oprint(self):
        if SvF.Compile :  print('Oprint Compile', self.Otype, self.name)
        else :            print('Oprint', self.Otype, self.name, "NoC", self.NoC, "NoR", self.NoR)

    def AddField( self, name, pos = -1 ) :
        if pos == -1:
            self.Flds.append ( Field ( name, '' ) )
            self.Flds[-1].tb = zeros ( (self.NoR), float64 )
        else :
            self.Flds.insert ( pos, Field ( name, '' ) )
            self.Flds[pos].tb = zeros ( (self.NoR), float64 )


    def Evaluate ( self, evalFld, byFld, byVal ) :   # оесортировано
        evalFldPoi = self.getField (evalFld)
        byFldPoi   = self.getField (byFld)
        search_beg = 0
        search_end = self.NoR-1
        if byVal <= byFldPoi.tb[search_beg]: return evalFldPoi.tb[search_beg]
        if byVal >= byFldPoi.tb[search_end]: return evalFldPoi.tb[search_end]
        i = searchsorted( byFldPoi.tb, byVal, side='left' )      # number
        prop = ( byVal - byFldPoi.tb[i-1] ) / ( byFldPoi.tb[i] - byFldPoi.tb[i-1] )
#        print ("iiiiiii", i, prop)
        value = evalFldPoi.tb[i-1] * (1-prop) + evalFldPoi.tb[i] * prop
        return value

    def IndexCol( self, Name ) :
        return self.getFieldNum(Name)

    def getFieldNum (self, name) :
        for ifi, ofi in enumerate (self.Flds) :
            if ofi.name == name:  return ifi
        return -1

    def getField (self, name) :
        for ofi in self.Flds:
            if ofi.name == name:   return ofi
        return None

    def getField_tb (self, name) :
        for ofi in self.Flds:
            if ofi.name == name:   return ofi.tb
        return None

    def dat (self, name) :              #  the same
        for ofi in self.Flds:
            if ofi.name == name:   return ofi.tb
        return None


    def KillField (self, name) :
                ind = self.getFieldNum (name) 
                if ind >= 0 : del  self.Flds[ind]

    def KillRow (self, num) :
        for ofi in self.Flds:
           ofi.tb = delete(ofi.tb, num)
        self.NoR -= 1
        self.sR = range (self.NoR)

    def KillRowsByMask (self, x_name, y_name, Mask, MaskVal) :
        xind = self.getFieldNum (x_name)
        yind = self.getFieldNum (y_name)
        NoR = 0
        for i in self.sR:
            mask_i = Mask.A[0].getPointNum (self.Flds[xind].tb[i])
            mask_j = Mask.A[1].getPointNum (self.Flds[yind].tb[i])
            if Mask.grd[mask_i,mask_j] == MaskVal : continue
            for ofi in self.Flds:
                ofi.tb[NoR] = ofi.tb[i]
            NoR += 1
        for ofi in self.Flds:
            ofi.tb = resize(ofi.tb, NoR)
        self.NoR = NoR
        self.sR = range(self.NoR)

    def AppendRec(self, *Parts):   # parts of Rec   ? что-то странное
            for i_part, part in enumerate(Parts):
#                print (i_part, part, self.Flds[i_part].tb)
                self.Flds[i_part].tb = append (self.Flds[i_part].tb, part)
#                print(i_part, part, self.Flds[i_part].tb)
            self.NoR += 1
            self.sR = range (self.NoR)

    def AppendRec1(self, Vals):  # последовательно список значений для новой записи
        for i_part, part in enumerate(Vals):
            #                print (i_part, part, self.Flds[i_part].tb)
            self.Flds[i_part].tb = append(self.Flds[i_part].tb, part)
        #                print(i_part, part, self.Flds[i_part].tb)
        self.NoR += 1
        self.sR = range(self.NoR)

    def Operation ( self, buf ) :
        part = buf.replace ( ' ', '').split ( '=' )
        rightPart = part[1]
        leftCol = part[0].split('.')[1]
        leftInd = self.getFieldNum(leftCol)
        for ic, c in enumerate(self.Flds) :
            rightPart = rightPart.replace('curentTabl.'+c.name, 'c.tb[:]')
        self.Flds[leftInd].tb[:] = eval(rightPart)

    def TblLonLatToGaussKruger ( self ) :
   #         print ('col 0 and 1 convert To GaussKruger')
#            print 'degree :', self.tbl[0][0], self.tbl[0][1]
            for i in range(self.NoR):
                self.Flds[0].tb[i], self.Flds[1].tb[i] = WGS84toGausKru(self.Flds[1].tb[i], self.Flds[0].tb[i], 0)
###            for i in range(self.tbl.shape[0]):
   ###             self.tbl[i][0], self.tbl[i][1] = WGS84toGausKru(self.tbl[i][1], self.tbl[i][0], 0)
#            print 'GaussKruger :', self.tbl[0][0], self.tbl[0][1]

    def where_con_list (self) :          #    создает список номеров аргументов
        if self.where_condition is None:  return
        con_list = []
        for a in self.fields_str.split(','):
 #          print ('fields_str', a)
            if a.strip() == '*': continue
            part = SplitIgnor(a, ' As ')
            if len(part) == 1:   args = part[0]
            else:                args = part[1]
    #        print (args)
            fn = self.getFieldNum( args)
            con_list.append (fn)
 #           print ('arrr', args, fn)
        self.con_list = con_list


    def CheckWhere (self, NoR) :       #   проверка  where
        if self.where_condition is None: return True
        args = []
        for ff in self.con_list : args.append (self.Flds[ff].tb[NoR])
        t_args = tuple(args)
        return self.where_condition (*t_args)



    def Read30_XLSX ( self ) :
        try :
            wb = openpyxl.load_workbook(self.fromFile)
            ws = wb[wb.get_sheet_names()[0]]
        except :
            print ('No file ', self.fromFile)
            exit (-1)
        else:
            print ('max column/row', ws.max_column, ws.max_row,)
            for nameRow in range (1,1000):           # Looking for NANEs string
                if ws.cell(row=nameRow,column=1).value == None : continue
                if str(ws.cell(row=nameRow,column=1).value)[0] == '#' : continue
                break
            print ('nameRow', nameRow)
            names = []
            for col in range(1,ws.max_column+1) :           # co - number of collomns - till first None
                name = str (ws.cell(row=nameRow,column=col).value)
                if name == '#END#' : break
                if name == None : name = ''
#                print name
                names.append(name.strip())           # remuve end  blancs
#            NoC = len (names)
            if SvF.printL :  print ("TablesNames", names)
 ##           print ("TablesNames", self.fromFile, ":", names)

            self.setField_src_num(names)
##            for fld in self.Flds : fld.Mprint()

#            for fld in self.Flds :
 #               if isnan(fld.src_num) :              # not  ROWNUM
  #                try :
   #                 fld.src_num = names.index ( fld.src_name )
    #              except :
     #               print ("No Column for", fld.src_name, "*****************************")
      #              exit (-1)
       #         fld.Mprint()

            self.where_con_list()       #   30

            self.NoC = len(self.Flds)  #(self.cols)
            maxNoR = ws.max_row  #50000
            for fld in self.Flds : fld.tb = zeros ( maxNoR, float64 ) 

            NoR = 0
            ro = nameRow
            for t_raw in range (ws.max_row-nameRow)  :  #while (1) :
                ro += 1
                try :
                    if str(ws.cell(row=ro,column=1).value) == '#END#' :
                        print ('#END#', NoR)
                        break
                    if str(ws.cell(row=ro,column=1).value)[0] == '#'  : continue
                except ValueError:  pass
                OK = True
                AllNaN = True
                for fld in self.Flds :
  #                  print (pceil.value, pceil.data_type )
                    if fld.src_num == -1 :
                        fld.tb[NoR] = NoR
                        continue
                    pceil = ws.cell(row=ro, column=fld.src_num + 1)
                    if SvF.TaskName[:5] == 'COVID' and fld.name == 'iso_code' :     ############################  COVID
                        fld.tb[NoR] = strTOnum(str(pceil.value))
                    elif SvF.TaskName[:8] == 'COVID-RW' and fld.name == 'date' :     ############################  COVID
                        fld.tb[NoR] = float(str(pceil.value).replace('-',''))
                    else:
                        fld.tb[NoR] = floatGradNaN(ws.cell(row=ro,column=fld.src_num+1).value)
                    if not self.useNaN and isnan(fld.tb[NoR]) : OK = False; break    #continue
                    if not isnan(fld.tb[NoR]):  AllNaN = False                 #2024.3.10
                if AllNaN : continue                                            #2024.3.10
                if not OK: continue
                if self.CheckWhere ( NoR) == False : continue    #   30
#                if not where is None:
 #                   args = []
  #                  for ff in con_list : args.append (self.Flds[ff].tb[NoR])
   #                 t_args = tuple(args)
    #                if where (*t_args) == False : continue
                NoR += 1
                if NoR >= maxNoR :                          # resize
                    maxNoR *= 2
   #                 print ('resize to',  maxNoR)
                    for fld in self.Flds : fld.tb = resize (fld.tb, maxNoR ) 
            for fld in self.Flds : fld.tb = resize (fld.tb, NoR ) 
            
            self.NoR = NoR
            print (NoR)

    def setField_src_num( self, names ):        # ИСПОЛЬЗОВАТЬ В ДР ФУНКЦИЯХ !
            for fld in self.Flds:
                if fld.src_name == '*':
                    for na in names: self.Flds.append(Field(na, na))
                    self.Flds.append(Field('ROWNUM', 'ROWNUM'))
                elif fld.src_name == 'ROWNUM':   fld.src_num = -1
                else:
                    fld.src_num = names.index(fld.src_name)
                    if fld.src_num == -1:
                        print("No Column for", fld.src_name, "*****************************")
                        exit(-1)
            self.KillField('*')
            self.NoC = len(self.Flds)

    def Read28_FUN(self) :
            srcFun = getFun ( self.fromFile )
            names = []                                      # имена аргументов и функции
            for a in srcFun.A:   names.append ( a.name )
            names.append(srcFun.V.name)

            self.setField_src_num ( names )

 ##           for fld in self.Flds : fld.Mprint()
            srcNoR = 1
            for a in srcFun.A: srcNoR *= (a.Ub + 1)
            for fld in self.Flds : fld.tb = zeros ( srcNoR, float64 )

            self.where_con_list()   ############### 30

            NoR = 0
            if srcFun.dim == 1:
                Ax = srcFun.A[0]
#                print (Ax.name)
                for x in Ax.NodS:
                    OK = True
                    for fld in self.Flds :
                        if   fld.src_num==-1 :  fld.tb[NoR] = NoR
                        elif fld.src_num== 0 :  fld.tb[NoR] = Ax.min + Ax.step * x
                        else :
                            fld.tb[NoR] = srcFun.grdNaNreal(r)
                            if not self.useNaN and isnan(fld.tb[NoR]) : OK = False; break
                    if not OK: continue
                    if self.CheckWhere(NoR) == False: continue  ############### 30
#                    if len (where) >= 3 :                       #check  where
 #                       wher = where.replace('ROWNUM',str(NoR))
  #                      for fld in self.Flds :
   #                         wher = SubstitudeName ( wher, fld.name, str(fld.tb[NoR]) )
    #                    if eval (wher,{}) == False : continue
                    NoR += 1
            elif srcFun.dim == 2:
                Ax = srcFun.A[0]
                Ay = srcFun.A[1]
                for y in Ay.NodS:
                  for x in Ax.NodS:
                    OK = True
                    for fld in self.Flds :
                        if   fld.src_num==-1 :  fld.tb[NoR] = NoR
                        elif fld.src_num== 0 :  fld.tb[NoR] = Ax.min + Ax.step * x
                        elif fld.src_num== 1 :  fld.tb[NoR] = Ay.min + Ay.step * y
                        else :
                            fld.tb[NoR] = srcFun.grdNaNreal(x,y)
                            if not self.useNaN and isnan(fld.tb[NoR]) : OK = False; break
                    if not OK: continue
                    if self.CheckWhere(NoR) == False: continue  ############### 30
#                    if len (where) >= 3 :                       #check  where
 #                       wher = where.replace('ROWNUM',str(NoR))
  #                      for fld in self.Flds :
   #                         wher = SubstitudeName ( wher, fld.name, str(fld.tb[NoR]) )
    #                    if eval (wher,{}) == False : continue
                    NoR += 1
            for fld in self.Flds : fld.tb = resize (fld.tb, NoR )
            self.NoR = NoR


    def Read30_TBL( self ) :    #  tbl -> fld.tb
            source_tb = getTbl ( self.fromFile )
            names = []                                      # имена аргументов и функции
            for f in source_tb.Flds:   names.append ( f.name )
#            print ('source_tb.Flds:', names)
            self.setField_src_num ( names )
#            for fld in self.Flds : fld.Mprint()

            self.NoC = len(self.Flds)
            for fld in self.Flds : fld.tb = zeros ( source_tb.NoR, float64 ) 

            self.where_con_list()   ############### 30

            NoR = 0
            for r in range (source_tb.NoR) :
                OK = True
                for c, fld in enumerate(self.Flds) :
                    if fld.src_num==-1 :  fld.tb[NoR] = NoR
                    else:                 fld.tb[NoR] = source_tb.Flds[fld.src_num].tb[r]
                    if (not self.useNaN) and isnan(fld.tb[NoR]) : OK = False; break    #continue
                if not OK: continue
                if self.CheckWhere(NoR) == False: continue   ############### 30

                NoR += 1
            for fld in self.Flds : fld.tb = resize (fld.tb, NoR )
            self.NoR = NoR


    def Read21_KML(self, where ) :   
        srcFlds = []                                            # обманем - смастерим таблицу
        srcFlds.append ( Field ( 'X', 'X' ) );  srcFlds[0].tb = [] 
        srcFlds.append ( Field ( 'Y', 'Y' ) );  srcFlds[1].tb = []

        with open( self.fromFile, "r") as fi:
            for line in fi :
                if line.find('<coordinates>') != -1 : break
            for line in fi :
                if line.find('</coordinates>') != -1 : break
                points = line.split()
    #            print (points)
                for p in points :
                    xyz = p.split(',')
                    srcFlds[0].tb.append ( float(xyz[0]) )       
                    srcFlds[1].tb.append ( float(xyz[1]) )       
     #               print (xyz)
            srcNoR = len(srcFlds[0].tb)
#            source_tb = getTbl ( self.fromFile )
            for fld in self.Flds :
                if fld.src_name == '*' :
                    for s_fld in srcFlds : self.Flds.append ( Field ( s_fld.name, s_fld.name ) )
                elif fld.src_name == 'ROWNUM' : fld.src_num = -1
                else :    
#                    fld.src_num = source_tb.getFieldNum (fld.src_name)
                    for is_fld, s_fld in enumerate(srcFlds) :
                         if s_fld.name == fld.src_name :
                             fld.src_num = is_fld
                             break
                    if fld.src_num == -1 :
                        print ("No Column for", fld.src_name, "*****************************")
                        exit (-1)    
            self.KillField ('*')
##            for fld in self.Flds : fld.Mprint()


            self.NoC = len(self.Flds)
            for fld in self.Flds : fld.tb = zeros ( srcNoR, float64 ) 

            NoR = 0;
            for r in range (srcNoR) :
                wher = where[0:]
                OK = True
                for c, fld in enumerate(self.Flds) :
                    if fld.src_num==-1 :  fld.tb[NoR] = NoR
                    else:                 fld.tb[NoR] = srcFlds[fld.src_num].tb[r]
                    if not self.useNaN and isnan(fld.tb[NoR]) : OK = False; break    #continue
                if not OK: continue
                if len (where) >= 3 :                       #check  where
                    wher = where.replace('ROWNUM',str(NoR))
                    for c , fld in enumerate(self.Flds) :
                        wher = SubstitudeName ( wher, fld.name, str(fld.tb[NoR]) )
                    if eval (wher,{}) == False : continue    
                NoR += 1
            for fld in self.Flds : fld.tb = resize (fld.tb, NoR )
            self.NoR = NoR


    def Read30_TXT( self ) :   #    tbl -> fld.tb  №№№ !!! С кодировкой Видоуз не работает - съедает первую строку
        if SvF.printL: print ('Read25_TXT')
        with open( self.fromFile, "r") as fi:
            line1 = fi.readline().strip('\ufeff\n')    # какая-то кодировка после Видоуз
#            print ('!'+line1+'!')
            names_ver_type = line1.split('#SvFver_')      # Version & Type of File
  #          print (names_ver_type)
            if len (names_ver_type) ==2 :
                parts = names_ver_type[1].split('_')
                self.FileVer  = int(parts[0])
                self.FileType = parts[1].split()[0]       # убрать окон. строки
            if SvF.printL : print ("FileVerType", self.FileVer, self.FileType)

            names = names_ver_type[0].split()
            if self.FileVer==0 and isfloat(names[0]) :       # if это число а не имя - таблица без назв столбцов
                names = ['Col'+str(c) for c in range(len(names)) ]
                fi.seek(0)

            if SvF.printL :  print ("TablesNames", names)
  ##              for fld in self.Flds :  fld.Mprint()

            for fld in self.Flds :
                if fld.src_name == '*' :
                    for na in names : self.Flds.append ( Field ( na, na ) )
                    self.Flds.append ( Field ( 'ROWNUM', 'ROWNUM' ) )
                elif fld.src_name == 'ROWNUM' : fld.src_num = -1
                else :    
                    fld.src_num = names.index(fld.src_name)
                    if fld.src_num == -1 :
                        print ("No Column for", fld.src_name, "*****************************")
                        exit (-1)    
            self.KillField ('*')
##            for fld in self.Flds : fld.Mprint()

            self.NoC = len(self.Flds)
            maxNoR = 50000;
            for fld in self.Flds : fld.tb = zeros ( maxNoR, float64 ) 

            self.where_con_list()       # 30

            NoR = 0;
            if self.FileType != 'mtr2' and self.FileType != 'matr2' :                   #  tbl
              for line in fi :
                nums_row = line.split()
           #     print ( 'F', nums_row[0] )
                if len(nums_row) == 0: continue
                if nums_row[0] == '#END#':  break
                if nums_row[0][0] == '#'    :  continue
                OK = True
                for fld in self.Flds :
                    if fld.src_num == -1 :   fld.tb[NoR] = NoR
                    else:
                        fld.tb[NoR] = floatGradNaN( nums_row[fld.src_num] )
                        if not self.useNaN and isnan(fld.tb[NoR]) : OK = False; break    #continue
                if not OK: continue
                if self.CheckWhere(NoR) == False: continue      # 30

                NoR += 1
                if NoR >= maxNoR :                          # resize
                    maxNoR *= 2
#                    print ('resize to',  maxNoR)
                    for fld in self.Flds : fld.tb = resize (fld.tb, maxNoR ) 
            else :                                                          # == 'matr2'
              XX = fi.readline().split()                      # x1
#              print len(XX), XX
              for line in fi :
                nums_row = line.split()
                if len (nums_row) == 0: continue
                wher = where[0:]
                for col, num in enumerate(nums_row) :
                    OK = True
                    if col == 0 : Y = float (num);  continue

                    for fld in self.Flds :
                        if   fld.src_num == -1 :   fld.tb[NoR] = NoR
                        elif fld.src_num ==  0 :   fld.tb[NoR] = float( XX[col-1] )  
                        elif fld.src_num ==  1 :   fld.tb[NoR] = Y
                        else:  # fld.src_num ==  2
                            try :
                                fld.tb[NoR] = float ( num )
                            except :
                                fld.tb[NoR] = NaN
                            if not self.useNaN and isnan(fld.tb[NoR]) : OK = False; break    #continue
                    if not OK: continue
                    if len (where) >= 3 :                       #check  where
                      wher = where.replace('ROWNUM',str(NoR))
                      for fld in self.Flds:
                        wher = SubstitudeName ( wher, fld.name, str(fld.tb[NoR]) )
 #                       for c in range(self.NoC):
    #                        wher = SubstitudeName(wher, self.cols[c], str(tbl[NoR, c]))

                      if eval (wher,{}) == False : continue
                    NoR += 1
                    if NoR >= maxNoR :                          # resize
                        maxNoR *= 2
 #                       print ('resize to',  maxNoR)
                        for fld in self.Flds : fld.tb = resize (fld.tb, maxNoR ) 
            for fld in self.Flds : fld.tb = resize (fld.tb, NoR ) 
#            print self.Flds[2].tb
            self.NoR = NoR
                
    def Read27_ASC( self, where ) :            # == 'matr Grid'               #    tbl -> fld.tb
        with open( self.fromFile, "r") as fi:
            names = ['X','Y','Z']
            if SvF.printL :  print ("TablesNames", names)
            for c in self.Flds :
                try :
#                    col_num = names.index(c.src_name) #[0])
                    c.src_num = names.index(c.src_name) #[0])
                except :    
                        print ("No Column for", c.src_name, "*****************************")
                        return    
##            for fld in self.Flds : fld.Mprint()


            grdX      = int(fi.readline().split()[1])
            grdY      = int(fi.readline().split()[1])
            XLLCORNER = float(fi.readline().split()[1])
            YLLCORNER = float(fi.readline().split()[1])
            CELLSIZE  = float(fi.readline().split()[1])
            NDT  = float(fi.readline().split()[1])
            if SvF.printL: print ("ReadGrig from", self.fromFile)  ###, self.cols  ###, self.nums
            if SvF.printL: print (grdX, grdY, XLLCORNER, YLLCORNER, CELLSIZE, NDT)

            x1 = zeros(grdX, float64)
            x2 = zeros(grdY, float64)
            for i in range(grdX): x1[i] = XLLCORNER + CELLSIZE * (i + 0.5)
            for j in range(grdY): x2[j] = YLLCORNER + CELLSIZE * (grdY-1 - j + 0.5)

###            self.NoC = len(self.cols)
            self.NoC = len(self.Flds)
            maxNoR = 50000;
            tbl = zeros ( (maxNoR, self.NoC), float64 )
            NoR = 0;

            r = where.split ('XYin')
            if len(r) == 2:  Rect = readListFloat19 (r[1])
            else          :  Rect = []
            if len (Rect) == 4 :
                  xmi = int ( ceil( (Rect[0] - XLLCORNER - CELLSIZE/2 ) / CELLSIZE - 1e-10 ) )    # !!!!!!!!!!!
                  xma = int (floor( (Rect[2] - XLLCORNER - CELLSIZE/2 ) / CELLSIZE + 1e-10 ) )    # !!!!!!!!!!!
                  yma = grdY - 1 - int ( ceil( (Rect[1] - YLLCORNER - CELLSIZE/2 ) / CELLSIZE - 1e-10 ))
                  ymi = grdY - 1 - int (floor( (Rect[3] - YLLCORNER - CELLSIZE/2 ) / CELLSIZE + 1e-10 ))
            else :
                  xmi = 0
                  xma = grdX - 1
                  yma = grdY - 1
                  ymi = 0
  #                print ("|||  xmi, xma, ymi, yma", xmi, xma, ymi, yma)
            
            for s in range(ymi) : fi.readline()
            line_num = ymi

            for s in range(yma-ymi+1):
                in_num = fi.readline().split()
                for col in range (len(x1)) :
                  OK = True                
                  if col < xmi : continue
                  if col > xma : break
                  for c in range(self.NoC) :
###                    if   self.nums[c]== 0 : tbl[NoR,c] = float(x1[col])
   ###                 elif self.nums[c]== 1 : tbl[NoR,c] = float(x2[line_num])
                    if   self.Flds[c].src_num== 0 : tbl[NoR,c] = float(x1[col])
                    elif self.Flds[c].src_num== 1 : tbl[NoR,c] = float(x2[line_num])
                    else :
                                            tbl[NoR,c] = float(in_num[col])
                                            if tbl[NoR,c] == NDT: tbl[NoR,c] = NaN
 #                                               in_num[col]= NaN
 #                                           tbl[NoR,c] = float(in_num[col])
                                            if not self.useNaN and isnan(tbl[NoR,c]) : OK = False; break   
                  if not OK: continue
#                  if len (where) >= 3 :                       #check  where
 #                   wher = where.replace('ROWNUM',str(NoR))   #  заменить на copy
  #                  for c in range(self.NoC) :
   #                     wher = SubstitudeName ( wher, self.cols[c], str(tbl[NoR,c]) )
    #                if eval (wher,{}) == False : continue    
                  NoR += 1
                  if NoR >= maxNoR :                          # resize
                    maxNoR *= 2
   #                 print ('resize to',  maxNoR)
                    tbl = resize (tbl, (maxNoR, self.NoC) )
                line_num += 1    
###            self.tbl = resize (tbl, (NoR, self.NoC) )
            tbl = resize (tbl, (NoR, self.NoC) )
            for ifld, fld in enumerate(self.Flds): fld.tb = tbl[:, ifld]  ############### Kill ######
#            for r in range (100) :
 #               print i, tbl[i,0], tbl[i,1], tbl[i,2]
            self.NoR = NoR
            return

    def WriteSvFtbl ( self, OutName = '', printL=0 ) :
        if OutName=='' :  OutName = self.name + '.txt'
        with open(OutName,'w') as f:  
            for fld in self.Flds :  f.write ( fld.name + '  ' )
            f.write ( '#SvFver_62_tbl' )
            for r in range ( self.NoR ) :
                f.write ('\n' )
                for fld in self.Flds :  f.write ('\t'+str(fld.tb[r]) )
        if printL : print ("Write to", OutName)

    def Old_Read19_ASC(self, _cols, where):  # == 'matr Grid'               #    tbl -> fld.tb
        with open(self.fromFile, "r") as fi:
            names = ['X', 'Y', 'Z']
            if SvF.printL:  print ("TablesNames", names)
            for c in _cols:
                try:
                    col_num = names.index(c[0])
                except:
                    print ("No Column for", c[0], "*****************************")
                    return
                self.nums.append(col_num)
                if len(c) == 3: c[0] = c[2]  ##  ????????????

            for i in range(len(_cols)):       self.cols.append(_cols[i][0])

            grdX = int(fi.readline().split()[1])
            grdY = int(fi.readline().split()[1])
            XLLCORNER = float(fi.readline().split()[1])
            YLLCORNER = float(fi.readline().split()[1])
            CELLSIZE = float(fi.readline().split()[1])
            NDT = float(fi.readline().split()[1])
            if SvF.printL: print ("ReadGrig from", self.fromFile, self.cols, self.nums)
            if SvF.printL: print (grdX, grdY, XLLCORNER, YLLCORNER, CELLSIZE, NDT)

            x1 = zeros(grdX, float64)
            x2 = zeros(grdY, float64)
            for i in range(grdX): x1[i] = XLLCORNER + CELLSIZE * (i + 0.5)
            for j in range(grdY): x2[j] = YLLCORNER + CELLSIZE * (grdY - 1 - j + 0.5)

            self.NoC = len(self.cols)
            maxNoR = 50000;
            tbl = zeros((maxNoR, self.NoC), float64)
            NoR = 0;

            r = where.split('XYin')
            if len(r) == 2:
                Rect = readListFloat19(r[1])
            else:
                Rect = []
            if len(Rect) == 4:
                xmi = int(ceil((Rect[0] - XLLCORNER - CELLSIZE / 2) / CELLSIZE - 1e-10))  # !!!!!!!!!!!
                xma = int(floor((Rect[2] - XLLCORNER - CELLSIZE / 2) / CELLSIZE + 1e-10))  # !!!!!!!!!!!
                yma = grdY - 1 - int(ceil((Rect[1] - YLLCORNER - CELLSIZE / 2) / CELLSIZE - 1e-10))
                ymi = grdY - 1 - int(floor((Rect[3] - YLLCORNER - CELLSIZE / 2) / CELLSIZE + 1e-10))
            else:
                xmi = 0
                xma = grdX - 1
                yma = grdY - 1
                ymi = 0
    #            print ("|||  xmi, xma, ymi, yma", xmi, xma, ymi, yma)

            for s in range(ymi): fi.readline()
            line_num = ymi

            for s in range(yma - ymi + 1):
                in_num = fi.readline().split()
                for col in range(len(x1)):
                    OK = True
                    if col < xmi: continue
                    if col > xma: break
                    for c in range(self.NoC):
                        if self.nums[c] == 0:
                            tbl[NoR, c] = float(x1[col])
                        elif self.nums[c] == 1:
                            tbl[NoR, c] = float(x2[line_num])
                        else:
                            tbl[NoR, c] = float(in_num[col])
                            if not self.useNaN and isnan(tbl[NoR, c]): OK = False; break
                    if not OK: continue
                    #                  if len (where) >= 3 :                       #check  where
                    #                   wher = where.replace('ROWNUM',str(NoR))   #  заменить на copy
                    #                  for c in range(self.NoC) :
                    #                     wher = SubstitudeName ( wher, self.cols[c], str(tbl[NoR,c]) )
                    #                if eval (wher,{}) == False : continue
                    NoR += 1
                    if NoR >= maxNoR:  # resize
                        maxNoR *= 2
     #                   print ('resize to', maxNoR)
                        tbl = resize(tbl, (maxNoR, self.NoC))
                line_num += 1
            self.tbl = resize(tbl, (NoR, self.NoC))
            #            self.tb_cols = names
            self.NoR = NoR
            return


