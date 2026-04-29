import numpy as np
import matplotlib.pyplot as plt
from scipy.integrate import solve_ivp
from scipy.optimize import root
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
# Dimensional Oregonator with additive random measurement noise
# and export of results to an Excel workbook (.xlsx)
#
# Model:
#   dX/dt = k1*A*Y - k2*X*Y + k3*A*X - 2*k4*X^2
#   dY/dt = -k1*A*Y - k2*X*Y + (k5/alpha)*Z
#   dZ/dt = 2*k3*A*X - k5*Z
#
# Variables:
#   X = [HBrO2]
#   Y = [Br-]
#   Z = [Ce4+]
#
# Noise model:
#   additive Gaussian noise on centered trajectories
#   X: sigma = 4 nM
#   Y: sigma = 0.6 nM
#   Z: sigma = 13 nM
# ============================================================

# Parameters
A = 0.016406707120152755       # mol/L
alpha = 2.0

k1 = 3.207070707070707         # L mol^-1 s^-1
k2 = 1.668100537200059e6       # L mol^-1 s^-1
k3 = 14.019255479322336        # L mol^-1 s^-1
k4 = 4.6415888336127726e4      # L mol^-1 s^-1
k5 = 0.12608268587175545       # s^-1

# Noise levels (standard deviations)
sigma_X_nM = 4.0
sigma_Y_nM = 0.6
sigma_Z_nM = 13.0

# Random seed for reproducibility
seed = 42
rng = np.random.default_rng(seed)

def rhs(t, u):
    X, Y, Z = u
    dXdt = k1 * A * Y - k2 * X * Y + k3 * A * X - 2.0 * k4 * X**2
    dYdt = -k1 * A * Y - k2 * X * Y + (k5 / alpha) * Z
    dZdt = 2.0 * k3 * A * X - k5 * Z
    return np.array([dXdt, dYdt, dZdt])

# Find equilibrium
guess = np.array([3.8e-7, 1.27e-7, 1.38e-6], dtype=float)
eq_sol = root(lambda u: rhs(0.0, u), guess, method="hybr")
if not eq_sol.success:
    raise RuntimeError("Equilibrium search failed.")
Xeq, Yeq, Zeq = eq_sol.x

# Initial condition near equilibrium
u0 = np.array([1.02 * Xeq, 0.98 * Yeq, 1.01 * Zeq], dtype=float)

# Integrate
t_span = (0.0, 5000.0)
t_eval = np.linspace(t_span[0], t_span[1], 20001)
#t_eval = np.linspace(t_span[0], t_span[1], 20000)
sol = solve_ivp(
    rhs,
    t_span,
    u0,
    method="BDF",
    t_eval=t_eval,
    rtol=1e-9,
    atol=1e-14,
)
if not sol.success:
    raise RuntimeError("ODE integration failed.")

t = sol.t
X, Y, Z = sol.y

# Deviations from equilibrium
dX = X #- Xeq
dY = Y #- Yeq
dZ = Z #- Zeq

# Established oscillatory regime only
#mask_tail = t >= 3500.0
mask_tail = t >= 2000.0
t_tail = t[mask_tail]
dX_tail = dX[mask_tail]
dY_tail = dY[mask_tail]
dZ_tail = dZ[mask_tail]

# Convert to nM
dX_nM = 1e9 * dX_tail
dY_nM = 1e9 * dY_tail
dZ_nM = 1e9 * dZ_tail

# Add Gaussian noise
dX_noisy_nM = dX_nM + rng.normal(0.0, sigma_X_nM, size=dX_nM.shape)
dY_noisy_nM = dY_nM + rng.normal(0.0, sigma_Y_nM, size=dY_nM.shape)
dZ_noisy_nM = dZ_nM + rng.normal(0.0, sigma_Z_nM, size=dZ_nM.shape)

# -------- Plotting --------
plt.figure(figsize=(8, 5))
plt.plot(t_tail, dX_nM, label="true")
plt.plot(t_tail, dX_noisy_nM, label="noisy", alpha=0.7)
plt.xlabel("time, s")
plt.ylabel("X - X*, nM")
plt.title("Dimensional Oregonator: X(t) with random errors")
plt.grid(True, alpha=0.3)
plt.legend()
plt.tight_layout()
plt.show()

plt.figure(figsize=(8, 5))
plt.plot(t_tail, dY_nM, label="true")
plt.plot(t_tail, dY_noisy_nM, label="noisy", alpha=0.7)
plt.xlabel("time, s")
plt.ylabel("Y - Y*, nM")
plt.title("Dimensional Oregonator: Y(t) with random errors")
plt.grid(True, alpha=0.3)
plt.legend()
plt.tight_layout()
plt.show()

plt.figure(figsize=(8, 5))
plt.plot(t_tail, dZ_nM, label="true")
plt.plot(t_tail, dZ_noisy_nM, label="noisy", alpha=0.7)
plt.xlabel("time, s")
plt.ylabel("Z - Z*, nM")
plt.title("Dimensional Oregonator: Z(t) with random errors")
plt.grid(True, alpha=0.3)
plt.legend()
plt.tight_layout()
plt.show()

plt.figure(figsize=(6, 6))
plt.plot(dX_nM, dZ_nM, label="true")
plt.plot(dX_noisy_nM, dZ_noisy_nM, label="noisy", alpha=0.6)
plt.xlabel("X - X*, nM")
plt.ylabel("Z - Z*, nM")
plt.title("Dimensional Oregonator: centered phase portrait with random errors")
plt.grid(True, alpha=0.3)
plt.legend()
plt.tight_layout()
plt.show()

# -------- Export to Excel --------
wb = Workbook()
ws = wb.active
ws.title = "Parameters"

header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
bold_font = Font(bold=True)

def style_header(row):
    for cell in row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

# Sheet 1: parameters / metadata
ws.append(["Name", "Value", "Units"])
style_header(ws[1])

rows = [
    ["A", A, "mol/L"],
    ["alpha", alpha, "-"],
    ["k1", k1, "L mol^-1 s^-1"],
    ["k2", k2, "L mol^-1 s^-1"],
    ["k3", k3, "L mol^-1 s^-1"],
    ["k4", k4, "L mol^-1 s^-1"],
    ["k5", k5, "s^-1"],
    ["sigma_X", sigma_X_nM, "nM"],
    ["sigma_Y", sigma_Y_nM, "nM"],
    ["sigma_Z", sigma_Z_nM, "nM"],
    ["seed", seed, "-"],
    ["Xeq", Xeq, "mol/L"],
    ["Yeq", Yeq, "mol/L"],
    ["Zeq", Zeq, "mol/L"],
    ["X0", u0[0], "mol/L"],
    ["Y0", u0[1], "mol/L"],
    ["Z0", u0[2], "mol/L"],
]
for r in rows:
    ws.append(r)

for col in ["A", "B", "C"]:
    ws.column_dimensions[col].width = 20

# Sheet 2: full raw solution
ws2 = wb.create_sheet("FullData")
ws2.append([
    "time_s",
    "X_mol_per_L", "Y_mol_per_L", "Z_mol_per_L",
    "dX_mol_per_L", "dY_mol_per_L", "dZ_mol_per_L"
])
style_header(ws2[1])

for ti, Xi, Yi, Zi, dXi, dYi, dZi in zip(t, X, Y, Z, dX, dY, dZ):
    ws2.append([float(ti), float(Xi), float(Yi), float(Zi), float(dXi), float(dYi), float(dZi)])

for col in ["A", "B", "C", "D", "E", "F", "G"]:
    ws2.column_dimensions[col].width = 18

# Sheet 3: tail region with true and noisy centered signals
ws3 = wb.create_sheet("TailCenteredNoisy")
ws3.append([
    "time_s",
    "dX_true_nM", "dY_true_nM", "dZ_true_nM",
    "dX_noisy_nM", "dY_noisy_nM", "dZ_noisy_nM"
])
style_header(ws3[1])

for vals in zip(t_tail, dX_nM, dY_nM, dZ_nM, dX_noisy_nM, dY_noisy_nM, dZ_noisy_nM):
    ws3.append([float(v) for v in vals])

for col in ["A", "B", "C", "D", "E", "F", "G"]:
    ws3.column_dimensions[col].width = 18

xlsx_name = "oregonator_results_noisy.xlsx"
wb.save(xlsx_name)
print(f"Saved Excel file: {xlsx_name}")
