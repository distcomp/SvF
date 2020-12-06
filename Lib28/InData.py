# -*- coding: cp1251 -*-

from  numpy import *

from  GaKru import *

import sys

import COMMON as com
#from   Table import *

from   GridVarArgs  import *

import openpyxl

#class Rectangle :                  #  Rectangle  ############

 #   def __init__ ( self ) :
  #      self.xMi = 0
   #     self.yMi = 0
    #    self.xMa = 0
     #   self.yMa = 0


def ReadPolygon ( InFile ):
      print ("\nInPolygonFile: ", InFile)
#      if   'XLSX' == self.InFile[-4:].upper() :
#            self.ReadXLSX()
#      else :                  # if 'TXT'  == self.InFile[-3:].upper() :
      try :
            fi = open ( InFile, "r")
      except IOError as e:
            print ("не удалось открыть файл", InFile)
            return None
      else :
        ret = []
        while (1):
           line = fi.readline().split()
           if len (line) == 0 : break
#           print len (line), line
           for l in line :
               xy = l.split(',')
#               print 'xy', xy
               ret.append (float(xy[0])) 
               ret.append (float(xy[1]))
        fi.close()
        return ret
        


def ReadGridInf ( ReadFrom, printL=0, Rect = [] ) :
      try :
            fi = open ( ReadFrom, "r")
      except IOError as e:
            print ("не удалось открыть файл", ReadFrom)
            return None, None, None, None;
      else :
              cols = ['X','Y','Z']
              grdX      = int(fi.readline().split()[1])
              grdY      = int(fi.readline().split()[1])
              XLLCORNER = float(fi.readline().split()[1])
              YLLCORNER = float(fi.readline().split()[1])
              CELLSIZE  = float(fi.readline().split()[1])
              NDT  = float(fi.readline().split()[1])
              if printL: print ("ReadSol from", ReadFrom, cols)
              if printL: print (grdX, grdY, XLLCORNER, YLLCORNER, CELLSIZE, NDT)
 #             print Rect
#              print (Rect[0] - XLLCORNER - CELLSIZE/2 ) / CELLSIZE - 1e-10
 #             print (Rect[2] - XLLCORNER - CELLSIZE/2 ) / CELLSIZE + 1e-10
  #            print (Rect[1] - YLLCORNER - CELLSIZE/2 ) / CELLSIZE - 1e-10
   #           print (Rect[3] - YLLCORNER - CELLSIZE/2 ) / CELLSIZE + 1e-10
              if len (Rect) == 4 :
                  xmi = int ( ceil( (Rect[0] - XLLCORNER - CELLSIZE/2 ) / CELLSIZE - 1e-10 ) )    # !!!!!!!!!!!
                  xma = int (floor( (Rect[2] - XLLCORNER - CELLSIZE/2 ) / CELLSIZE + 1e-10 ) )    # !!!!!!!!!!!
                  yma = grdY - 1 - int ( ceil( (Rect[1] - YLLCORNER - CELLSIZE/2 ) / CELLSIZE - 1e-10 ))
                  ymi = grdY - 1 - int (floor( (Rect[3] - YLLCORNER - CELLSIZE/2 ) / CELLSIZE + 1e-10 ))
                  print ("|||", xmi, xma, ymi, yma)
                  XLLCORNER += CELLSIZE * xmi
                  YLLCORNER += CELLSIZE * (grdY-1-yma)
                  grdX = xma-xmi+1
                  grdY = yma-ymi+1

                  gr = zeros((grdY, grdX),float64)
                  for s in range(ymi) : line = fi.readline()
                  for s in range(grdY):
                      line = fi.readline().split()
                      for c in range (grdX) :
                          gr[s][c] = float(line[c+xmi])
              else :
                  gr = loadtxt (fi,'double')

              for s in range(gr.shape[0]) : # grdY):
                  for c in range (gr.shape[1]): #grdX) :
                      if gr[s][c] == NDT : gr[s][c] = NaN

              gr_rev = zeros((grdY, grdX),float64)
              for s in range(grdY) : gr_rev[s][:] = gr[grdY-1-s][:]

              x1 = [] #zeros(grdX, float64)
              x2 = [] #zeros(grdY, float64)
              for i in range(grdX): x1.append( XLLCORNER + CELLSIZE * (i + 0.5) )
#              for i in range(grdX): x1[i] = XLLCORNER + CELLSIZE * (i + 0.5)
              for j in range(grdY): x2.append( YLLCORNER + CELLSIZE * (grdY-1 - j + 0.5) )
              x2 = sorted ( x2 )
              fi.close()
              if printL: print (gr_rev.shape)
      return cols, x1, x2, gr_rev


def ReadGridInftmp(ReadFrom, printL=0, Rect=[]):   # delete !
    try:
        fi = open(ReadFrom, "r")
    except IOError as e:
        print ("не удалось открыть файл", ReadFrom)
        return None, None, None, None;
    else:
        cols = ['X', 'Y', 'Z']
        grdX = int(fi.readline().split()[1])
        grdY = int(fi.readline().split()[1])
        XLLCORNER = float(fi.readline().split()[1])
        YLLCORNER = float(fi.readline().split()[1])
        CELLSIZE = float(fi.readline().split()[1])
        NDT = float(fi.readline().split()[1])
        if printL: print ("ReadSol from", ReadFrom, cols)
        if printL: print (grdX, grdY, XLLCORNER, YLLCORNER, CELLSIZE, NDT)
        #             print Rect
        #              print (Rect[0] - XLLCORNER - CELLSIZE/2 ) / CELLSIZE - 1e-10
        #             print (Rect[2] - XLLCORNER - CELLSIZE/2 ) / CELLSIZE + 1e-10
        #            print (Rect[1] - YLLCORNER - CELLSIZE/2 ) / CELLSIZE - 1e-10
        #           print (Rect[3] - YLLCORNER - CELLSIZE/2 ) / CELLSIZE + 1e-10
        if len(Rect) == 4:
            xmi = int(ceil((Rect[0] - XLLCORNER - CELLSIZE / 2) / CELLSIZE - 1e-10))  # !!!!!!!!!!!
            xma = int(floor((Rect[2] - XLLCORNER - CELLSIZE / 2) / CELLSIZE + 1e-10))  # !!!!!!!!!!!
            yma = grdY - 1 - int(ceil((Rect[1] - YLLCORNER - CELLSIZE / 2) / CELLSIZE - 1e-10))
            ymi = grdY - 1 - int(floor((Rect[3] - YLLCORNER - CELLSIZE / 2) / CELLSIZE + 1e-10))
            print ("|||", xmi, xma, ymi, yma)
            grdX = xma - xmi + 1
            grdY = yma - ymi + 1

            gr = zeros((grdY, grdX), float64)

            for s in range(ymi): line = fi.readline()
            for s in range(yma - ymi + 1):
                line = fi.readline().split()
                for c in range(xma - xmi + 1):
                    gr[s][c] = float(line[c + xmi])
            XLLCORNER += CELLSIZE * xmi
            YLLCORNER += CELLSIZE * (grdY - 1 - yma)
        else:
            gr = loadtxt(fi, 'double')
        #         print 'OOO', gr.shape
        #            1/0
        for s in range(gr.shape[0]):  # yma-ymi+1):
            for c in range(gr.shape[1]):  # xma-xmi+1) :
                if gr[s][c] == NDT: gr[s][c] = NaN

        x1 = zeros(grdX, float64)
        x2 = zeros(grdY, float64)
        for i in range(grdX): x1[i] = XLLCORNER + CELLSIZE * (i + 0.5)
        for j in range(grdY): x2[j] = YLLCORNER + CELLSIZE * (grdY - 1 - j + 0.5)
        fi.close()
        if printL: print (gr.shape)
    return cols, x1, x2, gr


def Get_Ver_Typ_cols ( head ):
            head = head.strip()
#            print head
            p_ver = head.find ('#SvFver_')
 #           print 'p_ver', p_ver
            if p_ver > 0 :
#                print 'head[p_ver:]', head[p_ver:]
                parts = head[p_ver+8:].split('_')
 #               print parts                 
                FileVer = int(parts[0])
                InFileTyp = parts[1]
                cols = head[0:p_ver].split()
#                print "RDcols: ", self.cols
                return FileVer, InFileTyp,  cols
            else :
                return 0, None, None



def ReadSolInf ( ReadFrom, printL=0 ) :
      try :
            fi = open ( ReadFrom, "r")
      except IOError as e:
            print ("не удалось открыть файл", ReadFrom)
            return None, None, None, None;
      else :
              if printL : print ("ReadSol from", ReadFrom)
              Ver,Typ,cols = Get_Ver_Typ_cols ( fi.readline() )
#              cols = fi.readline().split()
#              if printL : print "ReadSol from", ReadFrom,  cols,
              if not Ver > 0 :
                  print ("**************** Файл:", ReadFrom, 'не является решением')
                  exit (-1)
              if Typ != 'tbl' :                                   # 2-мерный
                  x_poi = fi.readline().split()
                  x1 = zeros( len(x_poi), float64 )
                  for j in range(len(x1)) : x1[j] = float(x_poi[j])
              else :                                                  # 1-мерный
                  x1  = []    #   ?????????????????
              tb = loadtxt (fi,'double')
              fi.close()
              if printL : print ("shape", tb.shape)
#              x2 = []
 #             for i in range(tb.shape[0]) : x2.append(tb[i,0])
              x2 = tb[:,0]
              if Typ == 'tbl':
                  grd = tb[:,1]
              else :
                  grd = delete(tb, range(0, 1), 1)
#              if Typ == 'tbl':
 #                 grd1 = zeros(tb.shape[0], float64)  #  Не умею убирать первое измерение
  #                grd1[:] = grd[:,0]
   #               grd = grd1
              #    grd = grd.transpose()

#      print 'q',x1,'f', x2
      return cols, x1, x2, grd




def SaveSolsXYTbl ( Sols, fName='', svf = '_SvF' ) :
    if fName == '' :  fName = "SumSolsTbl.tbl"
    fi = open ( fName, "w")
    Ax = Sols[0].A[0]
    Ay = Sols[0].A[1]
    fi.write ( Ax.name + '\t' + Ay.name )                   #  загол
    for s in Sols :  fi.write ( '\t' + s.V.name + svf )  #  загол
    for j in Ay.NodS :
        for i in Ax.NodS :
            fi.write ( "\n" + str(Ax.min + Ax.step*i) )     #  точки по х
            fi.write ( "\t" + str(Ay.min + Ay.step*j) )     #  точки по y
            for s in Sols : 
#                if not s.neNDT[i,j] :
 #                   print >> fi, "\t" + str (s.NDT),
  #              else :
                    if s.param :
                        print >> fi, "\t" + "%20.16g" % ((s.V.avr + s.grd[i,j]  )),   #  значения V
                    else :
                        if not s.neNDT[i, j]:
                            print >> fi, "\t" + str(s.NDT),
                        else:
                            print >> fi, "\t" + "%20.16g" % ((s.V.avr + s.grd[i,j]())),
    fi.close()
    print ("END of SaveSolsXYTbl")

def SaveSolsTbl ( Sols, fName='', svf = '_SvF' ) :
    if fName == '' :  fName = "SumSolsTbl.tbl"
    fi = open ( fName, "w")
    Ax = Sols[0].A[0]
    fi.write ( Ax.name )                   #  загол
    for s in Sols :  fi.write ( '\t' + s.V.name + svf )  #  загол
    for i in Ax.NodS :
            fi.write ( "\n" + str(Ax.min + Ax.step*i) )     #  точки по х
            for s in Sols :
                if not s.neNDT[i] :
                    print >> fi, "\t" + str (s.NDT),
                else :
                    if s.param :
                        print >> fi, "\t" + "%20.16g" % ((s.V.avr + s.grd[i]  )),   #  значения V
                    else :
                        print >> fi, "\t" + "%20.16g" % ((s.V.avr + s.grd[i]())),
    fi.close()
    print ("END of SaveSolsTbl")


class Cond :                 #  Cond   #########################################################
  def __init__ ( self, name, rel, Nval ) :     #  Имя, Условие, Число, Номер колонки имени
      self.name = name
      self.rel  = rel
      self.Nval = Nval
      self.num = -99

  def Cprint(self) :  print ("Cond", self.name, self.rel, self.Nval)  #, self.num


class Data :                  #  Data  ###################################  Data

  def __init__ ( self ) :
        self.InFile    = ""
        self.ColsToRead = []
        self.Npp       = 0          # Добавить № по прорядку
        self.File      = None
        self.InFileVer = 0
        self.InFileTyp = ''
        self.FileHead  = ""

        self.Cond = [ ]             # Cond
        self.Dtbl     = 0
        self.DNoC     = 0
        self.DNoR     = 0
        self.cols    = []
        self.NDT      = -99999
        self.mtr      = None
        self.mtrX1    = None
        self.mtrX2    = None
        self.TaskN    = ''
        self.Names    = ''
#        self.vNaN     = False
        self.ToGaussKruger = False
        self.Rect     = []

  def SubtitudeNames( self, shift=0 ) :     #  shift for Npp
      if self.Names == '' : return
#      self.cols = self.Names.split(',')
      nam = self.Names.split(',')
      for i, n in enumerate(nam) :
          if n != '': self.cols[i+shift] = n
      print (self.cols)

  def IndexCol( self, Name ) :
      if self.cols.count(Name) <= 0:
          print ("****IndexCol*********************** Err No Cols name for ***************", Name, "************")
          print ('Names', self.cols)
          return -99
      return self.cols.index(Name)

  def SetCondNum ( self ) :
      for c in self.Cond :
          c.num = self.IndexCol( c.name )
          print ('Cond_num', c.num,); c.Cprint()
          print (self.cols)
  def CheckCond ( self, i ) :
      for con in self.Cond:
#          print con.rel, self.Dtbl[i, con.num],  con.Nval
          if self.Dtbl[i, con.num] == self.NDT:                         return False
          if con.rel == "<"  and not self.Dtbl[i, con.num] <  con.Nval: return False
          if con.rel == ">"  and not self.Dtbl[i, con.num] >  con.Nval: return False
          if con.rel == "<=" and not self.Dtbl[i, con.num] <= con.Nval: return False
          if con.rel == ">=" and not self.Dtbl[i, con.num] >= con.Nval: return False
          if con.rel == "==" and not self.Dtbl[i, con.num] == con.Nval: return False
          if con.rel == "!=" and not self.Dtbl[i, con.num] != con.Nval: return False
      return True


  def ReadPolygon ( self ):
        print ("\nInPolygonFile: ", self.InFile)
#        if   'XLSX' == self.InFile[-4:].upper() :
#            self.ReadXLSX()
#        else :                  # if 'TXT'  == self.InFile[-3:].upper() :

        print (" Dtbl.shape ", self.Dtbl.shape)
        self.DNoC  = self.Dtbl.shape[1]
        self.DNoR  = self.Dtbl.shape[0]
        if self.ToGaussKruger: #Tbl_LatLonToGK(self.Dtbl)
            for i in range(self.Dtbl.shape[0]):
                self.Dtbl[i][0], self.Dtbl[i][1] = WGS84toGausKru(self.Dtbl[i][1], self.Dtbl[i][0], 0)
            print (self.Dtbl[0][0], self.Dtbl[0][1])

        if  self.InFile.count('Csz8_G-K.asc') == 1:
            print ('&&&&&&&&&&&&&&&&&  *37 For Blin @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@')
            self.Dtbl[:, 2] *= 37

  def TblToGaussKruger( self ) :
            for i in range(self.Dtbl.shape[0]):
                self.Dtbl[i][0], self.Dtbl[i][1] = WGS84toGausKru(self.Dtbl[i][1], self.Dtbl[i][0], 0)
            print (self.Dtbl[0][0], self.Dtbl[0][1])


  def Make_tbl ( self, x1, x2, mtr ):
      self.SubtitudeNames()
      print ("cols: ", self.cols, "mtr.shape", mtr.shape)
      print (len(x1), len(x2))
#      print x1.shape, x2.shape
      self.SetCondNum()
      if len(self.cols) == 2:  # 2
          self.Dtbl = zeros((mtr.shape[0], 2), float64)
          for i in range(self.Dtbl.shape[0]):
              self.Dtbl[self.DNoR, 0] = x1[i]
              self.Dtbl[self.DNoR, 1] = mtr[i]
              if self.CheckCond(self.DNoR):  self.DNoR += 1
      else:  # 3
          #            self.Dtbl = zeros( ( mtr.shape[0]*mtr.shape[1],3 ), float64 )
          self.Dtbl = zeros((min(10000, mtr.shape[0] * mtr.shape[1]), 3), float64)

          for x in range(mtr.shape[1]):
              for y in range(mtr.shape[0]):                                      # быстрый игрек
#          for y in range(mtr.shape[0]):
 #             for x in range(mtr.shape[1]):
                  self.Dtbl[self.DNoR, 0] = x1[x]
                  self.Dtbl[self.DNoR, 1] = x2[y]
                  self.Dtbl[self.DNoR, 2] = mtr[y, x]
                  if self.CheckCond(self.DNoR):
                      self.DNoR += 1
                      if self.DNoR >= self.Dtbl.shape[0]:
                          self.Dtbl = append(self.Dtbl, self.Dtbl, axis=0)
                          print ("append")
      self.Dtbl = delete(self.Dtbl, range(self.DNoR, self.Dtbl.shape[0]), 0)



  def ReadMatr2( self ) :
        x1 = self.File.readline().split()                      # x1
        self.mtrX1 = zeros( len(x1), float64 )
        for j in range(len(x1)) : self.mtrX1[j] = float(x1[j])

        self.mtr = loadtxt (self.File,'double')
        
        self.mtrX2 = zeros( self.mtr.shape[0], float64 )   # x2
        for i in range(len(self.mtrX2)) : self.mtrX2[i] = self.mtr[i,0]

        self.mtr = delete(self.mtr, range(0, 1), 1)
        print ("end of ReadMatr2")
        return
    
  def Matr2_To_Tbl( self, ReadDirection = 'X' ) :
        self.Dtbl = zeros( ( len (self.mtrX1)*len(self.mtrX2), 3 ), float64 )
        
        m = 0
        if 'X' == ReadDirection :
            for i in range(len(self.mtrX2)) : 
                for j in range(len(self.mtrX1)) :
                  self.Dtbl[m,0] = self.mtrX1[j]
                  self.Dtbl[m,1] = self.mtrX2[i]
                  self.Dtbl[m,2] = self.mtr[i,j]
                  m += 1
        else :  
            for j in range(len(self.mtrX1)) :
                for i in range(len(self.mtrX2)) : 
                  self.Dtbl[m,0] = self.mtrX1[j]
                  self.Dtbl[m,1] = self.mtrX2[i]
                  self.Dtbl[m,2] = self.mtr[i,j]
                  m += 1
                  
        if len(self.Cond) > 0 :
          self.SetCondNum()
          for r in range(self.Dtbl.shape[0]):
            if not self.CheckCond(r):  continue
            for c in range(self.Dtbl.shape[1]) :
                self.Dtbl[self.DNoR, c] = self.Dtbl[r, c]
            self.DNoR += 1
          self.Dtbl = delete(self.Dtbl, range(self.DNoR, self.Dtbl.shape[0]), 0)

        print ("end of Matr2_To_Tbl")


  def ReadData ( self ):
        if self.InFile == '' or self.DNoR != 0:  return
           
        with open(self.InFile,'r') as self.File:
            head = self.File.readline().strip()
#            print head
            ver = head.find ('#SvFver_')
            print ('ReadData ver', ver)
            if ver > 0 :
#                print 'head[ver:]', head[ver:]
                parts = head[ver+8:].split('_')
 #               print parts                 
                self.InFileVer = int(parts[0])
                self.InFileTyp = parts[1]
                self.InFileHead = head[0:ver]
                self.cols = self.InFileHead.split()
                self.SubtitudeNames()
                print ("RDcols: ", self.cols)

                if self.InFileTyp == 'matr2' :
                    self.ReadMatr2()
                    self.Matr2_To_Tbl()
                    self.mtr = None          # если сохранять, то нужно транспанировать 
                elif self.InFileTyp == 'tbl' :
                    self.ReadTXT (self.Rect)
                print (" Dtbl.shape ", self.Dtbl.shape)
#                if len(self.Dtbl.shape) == 0 :         #  1 number
 #                   self.DNoC  = 1
  #                  self.DNoR  = 1
   #             else :    
                self.DNoC  = self.Dtbl.shape[1]
                self.DNoR  = self.Dtbl.shape[0]
                return    
        print ("\nInFile: ", self.InFile, self.ColsToRead)

        if   'XLSX' == self.InFile[-4:].upper() :
            self.ReadXLSX18()
        elif 'SOL'  == self.InFile[-3:].upper() :
            self.cols, x1, x2, grd = ReadSolInf (self.InFile)
            self.Make_tbl( x1, x2, grd )
        elif 'ASC'  == self.InFile[-3:].upper() :
            self.cols, x1, x2, grd = ReadGridInf (self.InFile, 1, self.Rect)
            self.Make_tbl( x1, x2, grd )
        else :                  # if 'TXT'  == self.InFile[-3:].upper() :
            self.ReadTXT (self.Rect)
        print (" Dtbl.shape ", self.Dtbl.shape)
        self.DNoC  = self.Dtbl.shape[1]
        self.DNoR  = self.Dtbl.shape[0]
        if self.ToGaussKruger: #Tbl_LatLonToGK(self.Dtbl)
            for i in range(self.Dtbl.shape[0]):
                self.Dtbl[i][0], self.Dtbl[i][1] = WGS84toGausKru(self.Dtbl[i][1], self.Dtbl[i][0], 0)
            print (self.Dtbl[0][0], self.Dtbl[0][1])

        if  self.InFile.count('Csz8_G-K.asc') == 1:
            print ('&&&&&&&&&&&&&&&&&  *37 For Blin @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@')
            self.Dtbl[:, 2] *= 37

#        if self.TaskN == 'InUp' and self.InFile == 'X_Y_H.tbl' :
 #           self.Dtbl[:, 1] -= 65
  #          print '&&&&&&&&&&&&&&&&&  Shift For Blin @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@'
 #               self.Dtbl[:, 1] -= 50
  ###           self.Dtbl [:,0] += 30
  ###           self.Dtbl [:,1] += 80
  # ##      self.tbl [:,0] += 100     #  SRTM для Линника
  # ##      self.tbl [:,1] += 50



  def ReadXLSX18 ( self ):                ####  XLSX     устарело
        wb = openpyxl.load_workbook(self.InFile)
        ws = wb[wb.get_sheet_names()[0]]
        nameRo = 1
        for nameRo in range (1,1000):           # Looking for NANEs string
            if ws.cell(row=nameRo,column=1).value == None : continue
            if ws.cell(row=nameRo,column=1).value[0] == '#' : continue
            break
        for co in range(1,256) :                # co - number of collomns
            if ws.cell(row=nameRo,column=co).value == None : break
        for ro in range(nameRo,100000) :        # ro - num of rows
            if ws.cell(row=ro,column=1).value == None : break
        ro -= 2
        co -= 1
        print ('ro', ro, 'co', co, self.NDT)
#        self.cols = []
        for c in range(co) : self.cols.append(ws.cell(row=nameRo,column=c+1).value)
        print ("cols: ", self.cols)
        if len(self.ColsToRead) == 0 :                              #  без имен колонок
            self.SubtitudeNames()
            self.SetCondNum()
            self.Dtbl = zeros( ( ro,co ), float64 )        # без add   Npp
            for r in range(ro) :
              for c in range(co) :
                val = ws.cell(row=r+nameRo+1,column=c+1).value
                if val == None : self.Dtbl[self.DNoR,c] = self.NDT; continue
                if val == ' '  : self.Dtbl[self.DNoR,c] = self.NDT; continue
                if val == '  ' : self.Dtbl[self.DNoR,c] = self.NDT; continue
                if val == '   ': self.Dtbl[self.DNoR,c] = self.NDT; continue
                if val[0] == '#': self.Dtbl[self.DNoR,c] = self.NDT; continue
                self.Dtbl[self.DNoR,c] = val
              if self.CheckCond(self.DNoR) :  self.DNoR += 1
        else:                                                       #  с именами колонок

            self.Dtbl = zeros( ( ro,len(self.ColsToRead)+self.Npp ), float64 )        # add   Npp
            nums = []
            for name in self.ColsToRead : nums.append ( self.IndexCol( name ) )        # номера кол в файле
            print (nums)
            if self.Npp==1 : self.cols = ["Npp"]          # insert Npp
            else           : self.cols = []
            for name in self.ColsToRead : self.cols.append ( name )           # замена кол файла на кол таблицы
            self.SubtitudeNames(self.Npp)
            self.SetCondNum ()
            for r in range(ro) :
                if self.Npp==1 :  self.Dtbl[self.DNoR,0] = self.DNoR              # Npp
                OK = True
                for c in range(len(self.ColsToRead)) :
                    try :
                            self.Dtbl[self.DNoR,c+self.Npp] = float ( ws.cell(row=r+nameRo+1,column=nums[c]+1).value )
                    except :
                            self.Dtbl[self.DNoR,c+self.Npp] = NaN
                    if not com.useNaN and isnan(self.Dtbl[self.DNoR,c+self.Npp]) : OK = False; break   
                if not OK: continue
                if  self.CheckCond(self.DNoR) :   self.DNoR += 1
            print ('CCCC', self.cols, self.DNoR)
        self.Dtbl = delete(self.Dtbl, range(self.DNoR, self.Dtbl.shape[0]), 0)
        return


  def ReadTXT ( self, Rect=[] ):                     # TXT              
        if self.InFileVer == 0 :
            fi = open (self.InFile)
            self.cols = fi.readline().split()
            if '0123456789.'.count(self.cols[0][0]):
                for c in range (len(self.cols)) :
                    self.cols[c] = 'A' + str(c)
                fi.close()
                fi = open (self.InFile)
        else :
            fi = self.File
        self.SubtitudeNames()
        print ("cols: ", self.cols,)
        self.SetCondNum()
        if len(Rect) == 4:
            self.Dtbl = zeros((10000, len(self.cols)), float64)
            nr = 0
            while (1) :
                line = fi.readline()
                if line == '' : break
                line = line.split()
                for c in range(len(self.cols)) :
                    self.Dtbl[nr,c] = float (line[c])
                if  self.Dtbl[nr,0] < Rect[0] : continue
                if  self.Dtbl[nr,0] > Rect[2] : continue
                if  self.Dtbl[nr,1] < Rect[1] : continue
                if  self.Dtbl[nr,1] > Rect[3] : continue
                nr += 1
                if nr == self.Dtbl.shape[0]:
                    self.Dtbl = append(self.Dtbl, self.Dtbl, axis=0)
                    print ("append")
            self.Dtbl = delete(self.Dtbl, range(nr, self.Dtbl.shape[0]), 0)
        else:
            self.Dtbl = loadtxt (fi,'double')
        fi.close()
#        print 'type', type(self.Dtbl), len(self.Dtbl.shape) #dim
        if len(self.Dtbl.shape) == 0:              # 1 number
            tmp = zeros( ( 1,1 ), float64 )
            tmp[0,0] = self.Dtbl 
            self.Dtbl = tmp 
#            print '1_number', self.Dtbl, tmp.shape #dim
        for r in range(self.Dtbl.shape[0]):
            if not self.CheckCond(r):  continue
            for c in range(self.Dtbl.shape[1]) :
                self.Dtbl[self.DNoR, c] = self.Dtbl[r, c]
            self.DNoR += 1
            #        close
        self.Dtbl = delete(self.Dtbl, range(self.DNoR, self.Dtbl.shape[0]), 0)
        return

  # ##      self.tbl [:,0] += 100     #  SRTM для Линника
  # ##      self.tbl [:,1] += 50
  ###        if self.TaskN == 'InUp' :
  ###           self.Dtbl [:,0] += 30
  ###           self.Dtbl [:,1] += 80
  #           self.Dtbl [:,1] -= 50

  def WriteTbl ( self, OutFile ):
    fi = open (OutFile, "w")
    for co in self.cols : print >> fi, co, "\t",
    for r in range(self.Dtbl.shape[0]):
        print >> fi, '\n',
        for c in range(self.Dtbl.shape[1]):
            print >> fi, self.Dtbl[r,c], '\t',
    fi.close()











#class Cond :                 #  Cond   #########################################################
 # def __init__ ( self, name, rel, Nval ) :     #  Имя, Условие, Число, Номер колонки имени
  #    self.name = name
   #   self.rel  = rel
    #  self.Nval = Nval
     # self.Dnum = -99
#     # self.num  = -99
#  def Cprint(self) :  print "Cond", self.name, self.rel, self.Nval  #, self.num
 
class Col :                 #  допол колонки ####################################################
  def __init__ ( self, name ) :     #  Имя, Номер колонки имени
      self.name = name
      self.num  = -99
      self.Dnum = -99
  def Coprint(self) :  print ("DopCol", self.name, self.num)



#class InData (Data) :     ################################################################
class InData :     ################################################################
  def __init__ ( self ) :
#        Data.__init__(self)
    
        self.V = 0 #[ ]             #  var
        self.A = [ ]             #  Arg
        self.dim = -1
#        self.C = [ ]             # Cond
        self.Col = []
        self.tbl   = 0
        self.NoR   = 0
        self.sR    = []           # множ записей табл
        self.cols  = []
        self.NDT   = -99999
        self.SortedBy = Col('')
        self.mtr   = None 
        self.InFile = ''
        

  def Ar (self, name) :
      for a in range(len(self.A)) :
          if self.A[a].name == name : return self.A[a]
      return 0    
  

  def CheckNames ( self, DataR ):
    self.V.Dnum = DataR.IndexCol( self.V.name )

    Nnum = 0
    if self.V.Dnum == -99 :
        print ("***************************************** No NUM For Var", self.V.name, "****************")
    self.V.num = Nnum;  Nnum+=1;  self.cols.append(self.V.name)
#    print '\nDnum',self.V.Dnum, ; self.V.Vprint()

    for d in range( len(self.A) ) :
        self.A[d].Dnum = DataR.IndexCol(  self.A[d].name )
        if self.A[d].Dnum == -99 :
            print ("******************************* No NUM For Arg", self.A[d].name, "*****************************")
        self.A[d].num = Nnum;  Nnum+=1;  self.cols.append(self.A[d].name)
#        print 'Dnum',self.A[d].Dnum,; self.A[d].Aprint()

    for c in range(len(self.Col)) :
        self.Col[c].Dnum = DataR.IndexCol( self.Col[c].name )
        if self.Col[c].Dnum == -99 :
            print ("*************** Err No name for Col ***************", self.Col[c].Coprint())
        self.Col[c].num = Nnum;  Nnum+=1;  self.cols.append(self.Col[c].name)
#        print 'Dnum',self.Col[c].Dnum, ; self.Col[c].Coprint()

    if self.SortedBy.name != "" :
        self.SortedBy.Dnum = DataR.IndexCol( self.SortedBy.name )
        if self.SortedBy.Dnum == -99 :
            print ("************* Err No name for SortedBy ************", self.SortedBy.name, "************")
        self.SortedBy.num = Nnum;  Nnum+=1;  
#        print 'SortedBy', self.SortedBy.name, self.SortedBy.Dnum, self.SortedBy.num
    print ("End of CheckNames")


        
  def SubTbl ( self, DataR ) :
    self.InFile = DataR.InFile

    Nnum = 1 + len(self.A) + len(self.Col)
#    if self.SortedBy.Dnum > 0 :
 #       Nnum += 1 
    print ("SubTbl:  Nnum=", Nnum)
    Dtbl = DataR.Dtbl
    self.mtr = DataR.mtr              # for param

    self.tbl = zeros( ( Dtbl.shape[0],Nnum ), float64 )

    self.NoR = 0
    numNaN = 0
#    if str(type(self.tbl)) == '<type \'int\'>' : return
    for i in range(self.tbl.shape[0]) :
        if self.V.Dnum >= 0 :
            if isnan ( Dtbl[i,self.V.Dnum] ) :
                if com.useNaN == False: continue
                numNaN += 1 
            elif Dtbl[i,self.V.Dnum] == self.NDT : continue
        OK = 1
        for arg in self.A :
            if arg.Dnum < 0 : continue
            if isnan (Dtbl[i,arg.Dnum])     : OK = 0; break
            if Dtbl[i,arg.Dnum] == self.NDT : OK = 0; break
            if Dtbl[i,arg.Dnum] <  arg.min  : OK = 0; break 
            if Dtbl[i,arg.Dnum] >  arg.max  : OK = 0; break
            self.tbl[self.NoR,arg.num] = Dtbl[i,arg.Dnum]   # 1 ...
        if not OK : continue

        if self.V.Dnum >= 0 :
            self.tbl[self.NoR,self.V.num] = Dtbl[i,self.V.Dnum]                    # 0
#        for col in self.Col :  self.tbl[self.NoR,col.num] = Dtbl[i,col.Dnum]   #  ...
 #       if self.SortedBy.num > 0:
  #          self.tbl[self.NoR,self.SortedBy.num] = Dtbl[i,self.SortedBy.Dnum]   
        self.NoR += 1
    self.tbl = delete (self.tbl, range(self.NoR,self.tbl.shape[0]),0)

    if self.V.Dnum >= 0 :  self.V.dat = self.tbl[:,self.V.num]
    for arg in self.A :
            if arg.Dnum >= 0 : arg.dat = self.tbl[:,arg.num]
        
#    if self.SortedBy.num > 0 :
 #       def Comp (v1, v2) :
  #          if   v1[self.SortedBy.num] > v2[self.SortedBy.num] :   return  1
   #         elif v1[self.SortedBy.num] < v2[self.SortedBy.num] :   return -1
    #        return 0
     #   self.tbl = array (sorted(self.tbl, Comp))   # возрастание
    
    self.sR = range (self.NoR)    
#    print "numNaN", numNaN,
 #   print "NoR", self.NoR,
  #  print "mins", self.tbl.min(0)
   # print "maxs", self.tbl.max(0)


#  def ReadGrid_VS( self, ReadFrom, DataPath, ReadDirection = 'X' ) :
 #       if DataPath != '' : ReadFrom = DataPath + '/' + ReadFrom
  #      fi = open ( ReadFrom, "r")
   #     self.cols = fi.readline().split()
    #    print "cols: ", self.cols
     #   cols = fi.readline().split()
      #  x = zeros( len(cols), float64 )
       # for j in range(len(x)) : x[j] = float(cols[j])
#        tb = loadtxt (fi,'double')
 #       fi.close()
  #      print "Grid_VS", tb.shape
   #     self.tbl = zeros( ( len(x)*tb.shape[0],3 ), float64 )
    #    m = 0
     #   if 'X' == ReadDirection :
      #      for i in range(tb.shape[0]) : 
       #         for j in range(len(x)) :
        #          self.tbl[m,0] = x[j]
         #         self.tbl[m,1] = tb[i,0]
          #        self.tbl[m,2] = tb[i,j+1]
           #       m += 1
#        else :  
 #           for j in range(len(x)) :
  #              for i in range(tb.shape[0]) : 
   #               self.tbl[m,0] = x[j]
    #              self.tbl[m,1] = tb[i,0]
     #             self.tbl[m,2] = tb[i,j+1]
      #            m += 1
       # print "tbl.shape ", self.tbl.shape


  def ReadGrid( self, ReadGridFrom, DataPath ) :
        if DataPath != '' : ReadGridFrom = DataPath + '/' + ReadGridFrom
        fi = open ( ReadGridFrom, "r")
        grdX      = int(fi.readline().split()[1])
        grdY      = int(fi.readline().split()[1])
        XLLCORNER = float(fi.readline().split()[1])
        YLLCORNER = float(fi.readline().split()[1])
        CELLSIZE  = float(fi.readline().split()[1])
        self.NDT  = float(fi.readline().split()[1])
        print (grdX, grdY, XLLCORNER, YLLCORNER, CELLSIZE, self.NDT)
        gr = loadtxt (fi,'double')
        print (gr.shape)
        fi.close()

        self.tbl = zeros( (grdX*grdY,3), float64 )
        YSIZE = (grdY-1)*CELLSIZE;
        XLLCORNER += CELLSIZE*.5                            # на центр ячейки
        YLLCORNER += CELLSIZE*.5
        ma = 0;
        for y in range(grdY) :
          for x in range(grdX) :
            if gr[y,x] != self.NDT :
                    self.tbl[ma,0] = x * CELLSIZE + XLLCORNER
                    self.tbl[ma,1] = YSIZE - y * CELLSIZE +YLLCORNER
#                    self.tbl[ma,2] = gr[y,x] * 37
                    self.tbl[ma,2] = gr[y,x]
                    ma += 1
        print ("NoR", ma)
        self.tbl = delete (self.tbl, range(ma,self.tbl.shape[0]),0)
        if len(self.A)<1 :  self.A.append ( Arg('X') )
        if len(self.A)<2 :  self.A.append ( Arg('Y') )
        self.A[0].num = 0
        self.A[1].num = 1
#        self.V[0].num = 2
        self.V.num = 2
        self.dim  = 2



  def ToGrid ( self, OutFile, Gr ):
    A0 = self.A[0]
    A1 = self.A[1]
    f = open ( OutFile, "w")
    print  ("Write to ", OutFile)
    f.write( "NCOLS "   + str(A0.Ub+1) )
    f.write( "\nNROWS " + str(A1.Ub+1) )
    XLLCORNER = A0.min-A0.step*.5                   # на край ячейки
    YLLCORNER = A1.min-A1.step*.5                   # на край ячейки
    f.write( "\nXLLCORNER " + str(XLLCORNER) )
    f.write( "\nYLLCORNER " + str(YLLCORNER) )
    f.write( "\nCELLSIZE " + str(A0.step) )            # stepX !
    f.write( "\nNODATA_VALUE " + str(self.NDT) )
    for y in range(A1.Ub+1) :
         f.write( "\n" )
         for x in range(A0.Ub+1) :
             if Gr.grd1[ x,A1.Ub-y] == self.NDT :
                 f.write ( " " + str (self.NDT) )
             else :
                 #if MODE == ("N") then
#                 f.write ( " " + str( Gr.grd1[x, A1.Ub-y]() + self.V[0].avr ) )
                 f.write ( " " + str( Gr.grd1[x, A1.Ub-y]() + self.V.avr ) )
#                 f.write ( " 0" )
                 #if MODE == ("P") then  printf: " %f", (sum { px in s0pX, py in s0pY} c[px,py] * (x)^px * (grdY-y)^py) + V[0].avr > (OutFile); 
    f.close()
    print ("END of write into   ", OutFile)



  def SaveSol_D1 ( self, OutFile, Gr ):

    fi = open ( OutFile, "w" )
    fi.write ( self.A[0].name )
#    for v in range(len(self.V)) :
 #         fi.write ( "\t" + self.V[v].name + "_SvF")
    fi.write ( "\t" + self.V.name + "_SvF")
    for i in self.A[0].NodS :
        fi.write (   "\n" + str(self.A[0].min+self.A[0].step*i) 
                   + "\t" + str(Gr.grd1 [i]()*self.V.sig + self.V.avr) )
#                   + "\t" + str(Gr.grd1 [i]()*self.V[0].sig + self.V[0].avr) )
#                   + "\t" + str(Gr.grd [i]()) )
#        if len(self.V) > 1 :
 #         fi.write ( "\t" + str(Gr.grd2[i]() + self.V[1].avr) )
# #         fi.write ( "\t" + str(Gr.grd1[i]()) )
   #     if len(self.V) > 2 :
    #      fi.write (   "\t" + str(Gr.grd3[i]() + self.V[2].avr) )
                 
    fi.close()



        
  def SavePoints_D1 ( self, OutFile, Gr ):
    fi = open ( OutFile, "w" )
    fi.write ( self.A[0].name + "  " + self.V.name + "  " + self.V.name + "_SvFs\n" )
    for n in self.A[0].NodS :
        fi.write (   str(Gr.X[n]*self.A[0].step+self.A[0].min) + "  "
                   + str(Gr.F[n]+self.V.avr) + "  "
                   + str((Gr.F[n]-delta1(Gr,n)())+self.V.avr ) + "\n" )
    fi.close()

  def SaveCols(self) : 
        fi = open ( "SvF_"+self.A[0].name+"_Cols.txt", "w" )
        fi.write ( self.A[0].name )
        for c in range(len(self.Col)) :
            fi.write ( "\t" + self.Col[c].name)
        for n in range (self.NoR) :
            fi.write (   "\n" + str( self.A[0].min + self.A[0].step*self.tbl[n,self.A[0].num] ) )
            for c in range(len(self.Col)) :
                fi.write ( "\t" + str( self.tbl[n,self.Col[c].num] ) )
        fi.close()
        print ("Save Cols")






  def SigEst ( self, maxSigEst ) :
    sig2  = 0
    Nsig2 = 0
    for r in range(1,self.NoR) :
        if abs(self.tbl[r-1,self.A[0].num]-self.tbl[r,self.A[0].num]) <= maxSigEst :
                sig2  += ( 1/2.*self.tbl[r,self.V.num]-1/2.*self.tbl[r-1,self.V.num] )**2
                Nsig2 +=1
    sigSS2 = sqrt( 2 * sig2/Nsig2 )
    print ("Nsig2 = ", Nsig2, "     sigSS2 = ", sigSS2)
    sig3  = 0
    Nsig3 = 0
    for r in range(1,self.NoR-1) :
        if (    abs(self.tbl[r-1,self.A[0].num]-self.tbl[r,self.A[0].num]) <= maxSigEst
            and abs(self.tbl[r,self.A[0].num]-self.tbl[r+1,self.A[0].num]) <= maxSigEst ) :
                sig3  += (2/3.*self.tbl[r,self.V.num]-1/3.*self.tbl[r-1,self.V.num]-1/3.*self.tbl[r+1,self.V.num])**2
                Nsig3 +=1
    sigSS3 = sqrt( 3/2.* sig3/Nsig3 )
    print ("Nsig3 = ", Nsig3, "     sigSS3 = ", sigSS3)
    sig5  = 0
    Nsig5 = 0
    for r in range(2,self.NoR-2) :
        if (    abs(self.tbl[r-2,self.A[0].num]-self.tbl[r-1,self.A[0].num]) <= maxSigEst
            and abs(self.tbl[r-1,self.A[0].num]-self.tbl[r  ,self.A[0].num]) <= maxSigEst 
            and abs(self.tbl[r  ,self.A[0].num]-self.tbl[r+1,self.A[0].num]) <= maxSigEst 
            and abs(self.tbl[r+1,self.A[0].num]-self.tbl[r+2,self.A[0].num]) <= maxSigEst ) :
                sig5  += ( 4/5.*self.tbl[r,self.V.num]
                          -1/5.*self.tbl[r-2,self.V.num]-1/5.*self.tbl[r-1,self.V.num]
                          -1/5.*self.tbl[r+1,self.V.num]-1/5.*self.tbl[r+2,self.V.num])**2
                Nsig5 +=1
    sigSS5 = sqrt( 5/4.* sig3/Nsig5 )
    print ("Nsig5 = ", Nsig5, "     sigSS5 = ", sigSS5)




  def Normalization ( self, VarNormalization ) :
#      self.dim   = len(self.A)
#       print tbl.min(0)[self.num] 
      print (len(self.A))
      for arg in self.A :   arg.Normalization_UbSets ( )
      self.V.Normalization (VarNormalization) 
##      self.V.Normalization (self, VarNormalization) 


  def makeNodes ( self, n ) :    # n  -номер координаты
      print ('s', self.A[n].step)
      Xi = array ( floor( self.tbl[:,self.A[n].num] ),"int")
      grdX = int (ceil ( max (  self.tbl.max(0)[self.A[n].num],
                               (self.A[n].max-self.A[n].min)/self.A[n].step ) ) )
      for i in range(len(Xi)) :
          if Xi[i]==grdX : Xi[i]=grdX-1     # убирает данное с (не 0) границы.
      return Xi, grdX                       # вычисляем по внутренней четверке узлов



################################################################################################################



  def neNDT1 ( self, Ar1, Va ) :   # ---------  если хоть один из  VaS != NDT, то узел активен
      visA = Ar1.vis
      numA = Ar1.num

      neNDT = zeros ( Ar1.Ub+1,int8 )
#      neNDT = zeros ( Ar1.Ub+1,float64 )
      gap = ones ( Ar1.Ub+1,int8 )
      
      if visA==0 :
        for x in range( Ar1.Ub+1 ) :  neNDT[x] = 1
      else:
 # !!!!!!!!!!  разобраться зачем этот цикл !!!!!!!!!!!!!!!        
#        for v in range(len(VaS)) :
#          if VaS[v].num == -99 :  continue
        if Va.num != -99 : 
          for m in range(self.NoR) :
              if self.tbl[m,Va.num]!=self.NDT :
                  neNDT[floor(0.499999999+self.tbl[m,Ar1.num]/Ar1.step)] = 1 
        for m in range(self.NoR) :
          for x in range( max(0,           int(floor((self.tbl[m,numA]-visA)/Ar1.step))), 
                          min(Ar1.Ub,int(ceil ((self.tbl[m,numA]+visA)/Ar1.step)))+1 ) :
               neNDT[x] = 1
               
      Nxx = sum ( neNDT[x-1]*neNDT[x]*neNDT[x+1] for x in range( 1,Ar1.Ub ) )
#      print "Nxx.. ", Nxx
      return neNDT, Nxx, gap


#  NOT TESTED YET    
  def   neNDT2 (self, Ar1, Ar2, Va ) :   # ----------- если хоть один из  VaS != NDT, то узел активен

      neNDT = zeros ( (Ar1.Ub+1, Ar2.Ub+1), int8 )
#      neNDT = zeros ( (Ar1.Ub+1, Ar2.Ub+1), float64 )
      gap = ones ( (Ar1.Ub+1, Ar2.Ub+1),int8 )

      if Ar1.vis==0 and Ar2.vis==0  :
        for x in range( Ar1.Ub+1 ) :  
          for y in range( Ar2.Ub+1 ) :  neNDT[x,y] = 1
      else :
#        for m in range(self.NoR) :
#              if self.tbl[m,self.V[0].num]!=self.NDT :
#                  neNDT[int(self.tbl[m,Ar1.num]),int(self.tbl[m,Ar2.num])] = 1 
 # !!!!!!!!!!  разобраться зачем этот цикл !!!!!!!!!!!!!!!
###        for v in range(len(VaS)) :                 
###          if VaS[v].num == -99 :  continue
        if Va.num != -99 :  
          for m in range(self.NoR) :
#              if self.tbl[m,VaS[v].num]!=self.NDT :
#                  neNDT[int(self.tbl[m,Ar1.num]),int(self.tbl[m,Ar2.num])] = 1 
              if self.tbl[m,Va.num]!=self.NDT :
                  neNDT[floor(0.499999999+self.tbl[m,Ar1.num]/Ar1.step),
                        floor(0.499999999+self.tbl[m,Ar2.num]/Ar2.step)] = 1 

        for m in range(self.NoR) :
          for x in range( max(0,     int(floor((self.tbl[m,Ar1.num]-Ar1.vis)/Ar1.step))), 
                          min(Ar1.Ub,int(ceil ((self.tbl[m,Ar1.num]+Ar1.vis)/Ar1.step)))+1 ) :
            for y in range( max(0,     int(floor((self.tbl[m,Ar2.num]-Ar2.vis)/Ar2.step))), 
                            min(Ar2.Ub,int(ceil ((self.tbl[m,Ar2.num]+Ar2.vis)/Ar2.step)))+1 ) :
               neNDT[x,y] = 1

#      print Ar2.NodS,  "*********", Ar1.mNodSm
      Nxx = sum ( neNDT[x-1,y] * neNDT[x,y] * neNDT[x+1,y]  \
                    for y in Ar2.NodS    for x in Ar1.mNodSm )
      Nyy = sum ( neNDT[x,y-1] * neNDT[x,y] * neNDT[x,y+1]  \
                    for y in Ar2.mNodSm  for x in Ar1.NodS )
      Nxy = sum ( neNDT[x+1,y+1] * neNDT[x+1,y-1] * neNDT[x-1,y+1] * neNDT[x-1,y-1]
                    for y in Ar2.mNodSm  for x in Ar1.mNodSm )
#      print "Nxx.. ", Nxx, Nyy, Nxy
      return neNDT, Nxx, Nyy, Nxy, gap


      
